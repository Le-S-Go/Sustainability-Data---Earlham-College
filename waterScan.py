import pdfplumber
import pandas as pd
import calendar
import argparse
import os
from openpyxl import load_workbook

# my first try at writing a single bill parser, keep for now
def initial_single_bill(infile):
    with pdfplumber.open(infile) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text(keep_blank_chars=True)
    output = open("waterbilltxt.txt",'w')
    output.write(text)
    output = open('waterbilltxt.txt', 'r')
    lines = output.readlines()
    found_account = False # determine when the account number has been found
    for line in enumerate(lines):
        if 'Account No.' in line[1] and found_account == False:
            found_account = True
            account_number = line[1].split()[-1]
            account_number = account_number.split('-')[1]
            # account number has been found
        if 'ServicePeriod' in line[1]:
            month = identify_month(line[1])
        if 'Meter Reading and Usage Summary' in line[1]:
            finished = False # determine when the account information section has ended
            meter_dict = {}
            Count = 2
            while finished == False:
                next = lines[line[0]+Count].split()
                if 'A=Actual' in next:
                    finished = True
                    break
                meter = next[0]
                usage = next[-1].replace(',','')
                meter_dict.update({meter:[usage]})
                Count += 1
        if 'Total Service Related Charges' in line[1]:
            total_charge_line = line[1].split()
            total_charges = total_charge_line[4]
    total_usage = 0 # the total water usage across all meters on this account
    for meter in meter_dict.keys():
        total_usage += int(meter_dict.get(meter)[0])
    if total_usage == 0:
        for meter in meter_dict.keys():
            meter_charge = float(total_charges) / len(meter_dict.keys())
            meter_dict.get(meter).append(round(meter_charge,2))
    else:
        for meter in meter_dict.keys():
            proportion_used = int(meter_dict.get(meter)[0])/total_usage
            meter_charge = proportion_used * float(total_charges)
            meter_dict.get(meter).append(round(meter_charge,2))
    return(account_number, meter_dict, month)

### maybe for future work, not used now
month_dict = {'Jan':'Jan.', 'Feb':'Feb.', 'Mar':'Mar.', 'Apr':'Apr.', 'May':'May.', 'Jun':'Jun.', 'Jul':'Jul.', 'Aug':'Aug.', 'Sep':'Sept.', 'Oct':'Oct.', 'Nov':'Nov.', 'Dec':'Dec.'}       
def identify_month(monthline):
    line = monthline.split()
    month1 = line[1]
    date1 = line[2]
    month2 = line[4]
    date2 = line[5]
    if int(date1) < 15:
        correct_month = month_dict.get(month1)
    else:
        correct_month = month_dict.get(month2)
    return correct_month


def initialize_bill(infile):
    with pdfplumber.open(infile) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text(keep_blank_chars=True)
    output = open("waterbilltxt.txt",'w')
    output.write(text)
    output = open('waterbilltxt.txt', 'r')
    lines = output.readlines()
    weird_accounts = ['Account No.1010-210005414821\n', 'Account No.1010-210005483117\n','Account No.1010-210005543002\n', 'Account No.1010-210005543842\n', 'Account No.1010-210005865894\n', 'Account No.1010-210005543002\n'] 
    if 'Collective Invoice' in lines[1]: # this is a combined bill
        individual_dict = parse_combined_bill(lines)
        for key, value in individual_dict.items():
            account_dict.update({key:value})
    elif lines[2] in weird_accounts: # this is a weird single bill with no meter number that needs to be put in the account_dict
       # print('found weird bill')
        individual_dict = parse_weird_bill(lines)
        for key, value in individual_dict.items():
            account_dict.update({key:value})
    elif lines[2] == 'Account No.1010-210005789976\n': # this is the big bill
        #print('found big bill')
        individual_dict = parse_single_bill(lines, big_bill=True)
        for key, value in individual_dict.items():
            meter_dict.update({key:value})
    else:
        individual_dict = parse_single_bill(lines)
        for key, value in individual_dict.items():
            meter_dict.update({key:value})

def parse_weird_bill(lines):
    individual_dict = {}
    account_detail_line = lines[2].split()
    account_num = account_detail_line[1][8:]
    #print('found weird account num ' + account_num)
    for line in enumerate(lines):
        #if 'Account Detail' in line[1]: # beginning of account information section
         #   account_detail_line = lines[line[0]+1].split()
          #  account_num = account_detail_line[3][8:]
           # print(f'account number: {account_num}')
        if 'Total Current Period Charges' in line[1]: # this line contains the total charge for the account, which is all we have to work with for these weird bills
            total_charge_line = line[1].split()
            account_charge = total_charge_line[4]
    individual_dict.update({account_num:[0, account_charge]})
    return(individual_dict)

def parse_single_bill(lines, big_bill=False):
    first = True
    for line in enumerate(lines):
        if 'Meter Reading and Usage Summary' in line[1]: # beginning of account information section
            if big_bill == True and first == True:
                first = False
                continue
            finished = False # determine when the account information section has ended
            individual_dict = {} # stores meter number as key and a list of [usage, charge] as value for each meter on this account
            Count = 2
            while finished == False:
                next = lines[line[0]+Count].split()
                if 'A=Actual' in next:
                    finished = True
                    break
                meter = next[0]
                usage = next[-1].replace(',','')
                individual_dict.update({meter:[usage]})
                Count += 1
        if 'Total Current Period Charges' in line[1]:
            total_charge_line = line[1].split()
            total_charges = total_charge_line[4].replace(',','').replace('$','')
    total_usage = 0 # the total water usage across all meters on this account
    for meter in individual_dict.keys():
        total_usage += int(individual_dict.get(meter)[0])
    if total_usage == 0:
        for meter in individual_dict.keys():
            meter_charge = float(total_charges) / len(individual_dict.keys())
            individual_dict.get(meter).append(round(meter_charge,2))
    else:
        for meter in individual_dict.keys():
            proportion_used = int(individual_dict.get(meter)[0])/total_usage
            meter_charge = proportion_used * float(total_charges)
            individual_dict.get(meter).append(round(meter_charge,2))
    return(individual_dict)


def parse_combined_bill(lines):
    for line in enumerate(lines):
        if 'THANK YOU FOR BEING OUR CUSTOMER' in line[1]: # beginning of account information section
            finished = False # determine when the account information section has ended
            individual_dict = {} # stores account number as key and a list of [usage, charge] as value for each meter on this account
            counter = 1
            while finished == False:
                next = lines[line[0]+counter].split()
                if 'InvoiceTotals:' in next:
                    finished = True
                    break
                if len(next) > 1 and next[1] == 'EARLHAM': # this line contains the account number, amount, and usage information
                    account_num = next[0]
                    usage = next[4].replace(',','')
                    amount = next[6].replace(',','').replace('$','')
                    individual_dict.update({account_num:[usage, amount]})
                counter += 1
    return(individual_dict)

def update_excel(account_dict, meter_dict, blank_sheet, month):
    
    # Month order July → June
    month_order = [
        "Jul.", "Aug.", "Sept.", "Oct.", "Nov.", "Dec.",
        "Jan.", "Feb.", "Mar.", "Apr.", "May.", "Jun."]
    
    if month not in month_order:
        raise ValueError("Month must be abberviated month name (e.g., 'Jul.')")
    month_index = month_order.index(month)
    # Column math (Excel columns are 1-based in openpyxl)
    # Column G = 7
    base_column = 7
    cost_column = base_column + (month_index * 2)
    usage_column = cost_column + 1
    # Load workbook
    wb = load_workbook(blank_sheet)
    ws = wb["Water"]
    # Build lookup dictionaries for faster searching
    account_row_lookup = {}
    meter_row_lookup = {}
    for row in range(2, ws.max_row + 1):
        account_cell = ws.cell(row=row, column=5).value  # Column E
        meter_cell = ws.cell(row=row, column=6).value    # Column F
        
        if account_cell is not None:
            account_row_lookup[str(account_cell).strip()] = row
        
        if meter_cell is not None:
            meter_row_lookup[str(meter_cell).strip()] = row
    # Update account-based entries 
    for account, values in account_dict.items():
        if account in account_row_lookup:
            row = account_row_lookup[account]
            ws.cell(row=row, column=cost_column, value=float(values[1]))
            ws.cell(row=row, column=usage_column, value=float(values[0]))
        else:
            print(f"Warning: Account {account} not found in spreadsheet.")
    # Update meter-based entries 
    for meter, values in meter_dict.items():
        if meter in meter_row_lookup:
            row = meter_row_lookup[meter]
            ws.cell(row=row, column=cost_column, value=float(values[1]))
            ws.cell(row=row, column=usage_column, value=float(values[0]))
        else:
            print(f"Warning: Meter {meter} not found in spreadsheet.")
    # Save updated workbook
    wb.save(blank_sheet)
    print(f"Spreadsheet updated successfully for {month}.")      

if __name__ == "__main__":
    meter_dict = {}
    account_dict = {}
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", help="PDF files to be processed")
    parser.add_argument("-m", help="month") 
    parser.add_argument("-e", help='existing excel file')
    args = parser.parse_args()
    for file in os.listdir(args.f):
        infile = os.path.join(args.f, file)
        print(f"Processing {file}...")
        initialize_bill(infile)
    update_excel(account_dict, meter_dict, args.e, args.m)
    print('All files processed and spreadsheet updated successfully.')
    #for key, value in meter_dict.items():
    #    print(f'Meter: {key} Usage: {value[0]} Cost: ${value[1]}')
    #for key, value in account_dict.items():
     #   print(f'Account: {key} Usage: {value[0]} Cost: ${value[1]}')
